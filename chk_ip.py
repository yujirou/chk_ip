# ========================================================
# IP重複チェックバッチ
# ※チェック対象のエクセルを閉じていないとエラーになるので注意
# ========================================================
import os
import openpyxl as excel
import ipaddress as ip

# --------------------------------------------------------
# 設定
# --------------------------------------------------------
# チェック対象エクセル
file_path = "./list.xlsx"
# チェック対象エクセルのシート
work_sheet = "ip"
# チェック開始行
start_row = 3
# チェックデータ列（A列=0,B列=1,...）
ip_src_col = 0
ip_dst_col = 1
port_col = 2
# チェック結果出力列
info_col = "D"

# --------------------------------------------------------
# 処理
# --------------------------------------------------------
wb = excel.load_workbook(file_path)
ws = wb[work_sheet]

# 必要なデータを取得
data_list = []
for row in ws.iter_rows(min_row=start_row, min_col=1):
	tmp = []
	for i, cell in enumerate(row):
		if i in [ip_src_col, ip_dst_col, port_col]:
			tmp.append(cell.value)
	data_list.append(tmp)

# 重複チェック
info_1 = {} # 内包される側の結果を格納
info_2 = {} # 内包する側の結果を格納
for ai, ad in enumerate(data_list):
	a_src = ad[0]
	a_dst = ad[1]
	a_prt = ad[2]
	if a_src == None or a_dst== None or a_prt == None: continue
	a_src_ip = ip.ip_network(a_src)
	a_dst_ip = ip.ip_network(a_dst)
	for bi, bd in enumerate(data_list):
		if ai == bi: continue
		b_src = bd[0]
		b_dst = bd[1]
		b_prt = bd[2]
		if b_src == None or b_dst== None or b_prt == None: continue
		b_src_ip = ip.ip_network(b_src)
		b_dst_ip = ip.ip_network(b_dst)
		if a_prt == b_prt and a_dst_ip.subnet_of(b_dst_ip) and a_src_ip.subnet_of(b_src_ip):
			info_1[ai+start_row] = str(bi+start_row)+"行目に含まれています。"
			info_2[bi+start_row] = "内包している設定が存在します。"

print(info_1)
print(info_2)

# 情報出力用の列の値を削除
for row, cell in enumerate(ws[info_col]):
	if row == start_row-2:
		cell.value = "重複チェックバッチ出力用"
		continue
	cell.value = None

# 情報出力用の列にチェックした結果を出力
for i, v in info_1.items():
	ws[info_col+str(i)].value = v

for i, v in info_2.items():
	ws[info_col+str(i)].value = v


wb.save(file_path)

